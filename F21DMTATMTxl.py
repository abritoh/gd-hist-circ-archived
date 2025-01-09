# -*- coding: utf-8 -*-
"""
-- ============================================================================
-- Descripción....: 
	(1) Generación de Reporte Excel: 
        Formato 2.1 - Demandas mensuales de transformadores AT-MT
    
-- Elabora........: ArBR (arcebrito@gmail.com)
-- Fecha..........: 2019-02-05
-- ============================================================================
"""

import time
import utlpy
import shutil
import pymysql
import openpyxl
from openpyxl import load_workbook
from threading import Thread

ZERO_ROW            = 7
REPOXLS_PATH        = ""
PR_NAME             = "F21DMTATM"
LOGO_CFE_DIST_PATH  = "TEMPLATE\logo_cfe_distribucion.png"
TEMPLATE_PATH       = "TEMPLATE\F21DMTatmt_template.xlsx"
REPOXLS_BASE_PATH   =  'E:\\repoxls' if utlpy.drive_exists('e') else 'c:\\repoxls'

CN = {"host": "10.4.22.84", "user":  "ClusterBR", "passwd": "XYZ123...", "database": "apcc"}

def fn_get_ds_dem_transformadores_mes(connection, cveDivision, cveZona, anio) :
    qry = """
            SELECT 
                CZ.ABREVIATURA, BSUB, BNOMSUB, BBANCO, 
                BCAPACIDAD, CONCAT(BTENBAJA, "/", BTENALTA) AS RTENSION, BNUMCIR,
            	(SELECT BHDEM FROM APCC.DM_BANCOHIST H WHERE H.BHDIV=B.BDIV AND H.BHZONA=B.BZONA AND H.BHSUB=B.BSUB AND H.BHBAN=B.BBANCO AND H.BHULTACT='[ANIO]' AND H.BHMES='ENE' LIMIT 1) AS ENE,
            	(SELECT BHDEM FROM APCC.DM_BANCOHIST H WHERE H.BHDIV=B.BDIV AND H.BHZONA=B.BZONA AND H.BHSUB=B.BSUB AND H.BHBAN=B.BBANCO AND H.BHULTACT='[ANIO]' AND H.BHMES='FEB' LIMIT 1) AS FEB,
            	(SELECT BHDEM FROM APCC.DM_BANCOHIST H WHERE H.BHDIV=B.BDIV AND H.BHZONA=B.BZONA AND H.BHSUB=B.BSUB AND H.BHBAN=B.BBANCO AND H.BHULTACT='[ANIO]' AND H.BHMES='MAR' LIMIT 1) AS MAR,
            	(SELECT BHDEM FROM APCC.DM_BANCOHIST H WHERE H.BHDIV=B.BDIV AND H.BHZONA=B.BZONA AND H.BHSUB=B.BSUB AND H.BHBAN=B.BBANCO AND H.BHULTACT='[ANIO]' AND H.BHMES='ABR' LIMIT 1) AS ABR,
            	(SELECT BHDEM FROM APCC.DM_BANCOHIST H WHERE H.BHDIV=B.BDIV AND H.BHZONA=B.BZONA AND H.BHSUB=B.BSUB AND H.BHBAN=B.BBANCO AND H.BHULTACT='[ANIO]' AND H.BHMES='MAY' LIMIT 1) AS MAY,
            	(SELECT BHDEM FROM APCC.DM_BANCOHIST H WHERE H.BHDIV=B.BDIV AND H.BHZONA=B.BZONA AND H.BHSUB=B.BSUB AND H.BHBAN=B.BBANCO AND H.BHULTACT='[ANIO]' AND H.BHMES='JUN' LIMIT 1) AS JUN,
            	(SELECT BHDEM FROM APCC.DM_BANCOHIST H WHERE H.BHDIV=B.BDIV AND H.BHZONA=B.BZONA AND H.BHSUB=B.BSUB AND H.BHBAN=B.BBANCO AND H.BHULTACT='[ANIO]' AND H.BHMES='JUL' LIMIT 1) AS JUL,
            	(SELECT BHDEM FROM APCC.DM_BANCOHIST H WHERE H.BHDIV=B.BDIV AND H.BHZONA=B.BZONA AND H.BHSUB=B.BSUB AND H.BHBAN=B.BBANCO AND H.BHULTACT='[ANIO]' AND H.BHMES='AGO' LIMIT 1) AS AGO,
            	(SELECT BHDEM FROM APCC.DM_BANCOHIST H WHERE H.BHDIV=B.BDIV AND H.BHZONA=B.BZONA AND H.BHSUB=B.BSUB AND H.BHBAN=B.BBANCO AND H.BHULTACT='[ANIO]' AND H.BHMES='SEP' LIMIT 1) AS SEP,
            	(SELECT BHDEM FROM APCC.DM_BANCOHIST H WHERE H.BHDIV=B.BDIV AND H.BHZONA=B.BZONA AND H.BHSUB=B.BSUB AND H.BHBAN=B.BBANCO AND H.BHULTACT='[ANIO]' AND H.BHMES='OCT' LIMIT 1) AS OCT,
            	(SELECT BHDEM FROM APCC.DM_BANCOHIST H WHERE H.BHDIV=B.BDIV AND H.BHZONA=B.BZONA AND H.BHSUB=B.BSUB AND H.BHBAN=B.BBANCO AND H.BHULTACT='[ANIO]' AND H.BHMES='NOV' LIMIT 1) AS NOV,
            	(SELECT BHDEM FROM APCC.DM_BANCOHIST H WHERE H.BHDIV=B.BDIV AND H.BHZONA=B.BZONA AND H.BHSUB=B.BSUB AND H.BHBAN=B.BBANCO AND H.BHULTACT='[ANIO]' AND H.BHMES='DIC' LIMIT 1) AS DIC
            FROM APCC.BANCOS B
            INNER JOIN APCC.TBCATZONA CZ ON CZ.CLAVEDIVISION = B.BDIV AND CZ.CLAVE = BZONA
            WHERE B.BDIV = %S AND B.BZONA = %S
            ORDER BY CZ.ABREVIATURA, BSUB, BNOMSUB, BBANCO
            """    
    qry = qry.replace("[ANIO]", anio)
    params = (cveDivision, cveZona)
    return utlpy.mysql_fecthall_dict(connection, qry, params)

def insert_logo_cfe(ws) :    
    img = openpyxl.drawing.image.Image(LOGO_CFE_DIST_PATH)
    ws.cell(row=2, column=2)
    ws.add_image(img)
    return

def style_cell_medicion(cell):    
    utlpy.pyxl_center_borderall_fill_cell(cell, fill=False, center=True, border=True, fgColor="FFFFFF")  
    return

def fn_create_rpt_by_zona(connection, anio, cveDivision, wb, rs_zona) :
    
    cveZona = rs_zona["claveZona"]        
    ws = wb.copy_worksheet(wb["ZONA"])
    ws.title = "{}{}".format(cveDivision,cveZona)    
    
    insert_logo_cfe(ws)    
    nombre_division = utlpy.fn_get_division_only(rs_zona["NombreDivision"])
    cell_div = ws.cell(row = 2, column = 1)
    cell_zon = ws.cell(row = 3, column = 1)
    cell_anio = ws.cell(row = 5, column = 1)    
    cell_div.value = cell_div.value.replace("[NOMBRE_DIVISION]", nombre_division)
    cell_zon.value = cell_zon.value.replace("[NOMBRE_ZONA]", rs_zona["Titulo"])
    cell_anio.value = cell_anio.value.replace("[ANIO]", anio)
    
    idx_banco = 0
    ds_demanda_mes = fn_get_ds_dem_transformadores_mes(connection, cveDivision, cveZona, anio)
    for rs in ds_demanda_mes :
        idx_banco = idx_banco + 1
        utlpy.println("{}-{}-{} - Banco:{}".format(cveDivision, cveZona, rs["bnomsub"], rs["bbanco"]))
        
        style_cell_medicion(ws.cell(row = ZERO_ROW + idx_banco, column = 1, value = rs["Abreviatura"]))
        style_cell_medicion(ws.cell(row = ZERO_ROW + idx_banco, column = 2, value = rs["bsub"]))
        style_cell_medicion(ws.cell(row = ZERO_ROW + idx_banco, column = 3, value = rs["bnomsub"]))
        style_cell_medicion(ws.cell(row = ZERO_ROW + idx_banco, column = 4, value = utlpy.to_number(rs["bbanco"])))
        style_cell_medicion(ws.cell(row = ZERO_ROW + idx_banco, column = 5, value = utlpy.to_number(rs["bcapacidad"])))
        style_cell_medicion(ws.cell(row = ZERO_ROW + idx_banco, column = 6, value = rs["rtension"]))   
        style_cell_medicion(ws.cell(row = ZERO_ROW + idx_banco, column = 7, value = rs["bnumcir"]))   
        
        style_cell_medicion(ws.cell(row = ZERO_ROW + idx_banco, column = 8, value = rs["ENE"]))
        style_cell_medicion(ws.cell(row = ZERO_ROW + idx_banco, column = 9, value = rs["FEB"]))
        style_cell_medicion(ws.cell(row = ZERO_ROW + idx_banco, column = 10, value = rs["MAR"]))
        style_cell_medicion(ws.cell(row = ZERO_ROW + idx_banco, column = 11, value = rs["ABR"]))
        style_cell_medicion(ws.cell(row = ZERO_ROW + idx_banco, column = 12, value = rs["MAY"]))
        style_cell_medicion(ws.cell(row = ZERO_ROW + idx_banco, column = 13, value = rs["JUN"]))
        style_cell_medicion(ws.cell(row = ZERO_ROW + idx_banco, column = 14, value = rs["JUL"]))
        style_cell_medicion(ws.cell(row = ZERO_ROW + idx_banco, column = 15, value = rs["AGO"]))
        style_cell_medicion(ws.cell(row = ZERO_ROW + idx_banco, column = 16, value = rs["SEP"]))
        style_cell_medicion(ws.cell(row = ZERO_ROW + idx_banco, column = 17, value = rs["OCT"]))
        style_cell_medicion(ws.cell(row = ZERO_ROW + idx_banco, column = 18, value = rs["NOV"]))
        style_cell_medicion(ws.cell(row = ZERO_ROW + idx_banco, column = 19, value = rs["DIC"]))
    #end-for
    
    return

def proc_genera_xls_by_zona (anio, cveDivision, cveZona, rs_zona) :
    
    start_time = time.time()    
    prid = utlpy.btc_gen_prid(PR_NAME, "{}{}{}".format(cveDivision, cveZona, anio))
    connection = pymysql.connect(host=CN["host"], user=CN["user"], passwd=CN["passwd"], database=CN["database"]) 
    dest_fname = "{}\{}_{}{}{}.xlsx".format(REPOXLS_PATH, PR_NAME, anio, cveDivision, cveZona)
    
    utlpy.btc_insert(connection, prid, PR_NAME, "INICIADO", "", cveDivision, cveZona, "", "", anio, "*")
    try:
        utlpy.println("proc_genera_xls_by_zona:{} thread started & running (...)".format(cveZona))
        
        shutil.copy(TEMPLATE_PATH, dest_fname)
        wb = load_workbook(dest_fname)
        
        fn_create_rpt_by_zona(connection, anio, cveDivision, wb, rs_zona)
        
        wb.remove(wb["ZONA"])
        wb.save(dest_fname) 
        
        elapsed_time = time.time() - start_time
        elapsed_time_fmt = time.strftime("%H:%M:%S", time.gmtime(elapsed_time))
        
        message = "DURACION {}".format(str(elapsed_time_fmt))
        utlpy.println("proc_genera_xls_by_zona:{} thread completed. Total time: {}".format(cveZona, elapsed_time_fmt))
        utlpy.btc_insert(connection, prid, PR_NAME, "COMPLETADO", message, cveDivision, cveZona, "", "", anio, "*")
        
    except Exception as e:
        message = str(e)
        utlpy.println('proc_genera_xls_by_zona.error>>> ' + message)        
        utlpy.btc_insert(connection, prid, PR_NAME, "ERROR", message, cveDivision, cveZona, "", "", anio, "*")
        
    finally:
        connection.close()        

    return

def proc_genera_xls_by_div(cveDivision, anio) :    
    
    utlpy.println("proc_genera_xls_by_div: >>> main thread started >>>")
    
    connection = None
    try:
        connection = pymysql.connect(host=CN["host"], user=CN["user"], passwd=CN["passwd"], database=CN["database"]) 
        lst_threads = []
        ds_zonas = utlpy.fn_get_ds_zonas(connection, cveDivision)
        connection.close()
        connection = None
        
        for rs_zona in ds_zonas :
            cveZona = rs_zona["claveZona"]
            t = Thread(target = proc_genera_xls_by_zona, args = (anio, cveDivision, cveZona, rs_zona))
            lst_threads.append(t)
        
        [t.start() for t in lst_threads]
        [t.join() for t in lst_threads]
            
        zip_file_name = "{}_{}{}".format(PR_NAME, anio, cveDivision)
        utlpy.create_parent_zip_from_dir(REPOXLS_PATH, zip_file_name)
        
    except Exception as e:
        utlpy.println('proc_genera_xls_by_div.error>>> {}'.format(str(e)))
    finally:
        if connection:
            connection.close()
        
    print("proc_genera_xls_by_div: >>> main thread completed >>>")
    return


##########################
# main
##########################

if __name__ == '__main__' :
    
    anio = "2018"
    lst_divisiones = ['DA','DB','DC','DD','DF','DG','DJ','DK','DL','DM','DN','DP','DU','DV','DW','DX']
    for division in lst_divisiones :
        REPOXLS_PATH = "{}\{}\{}\{}".format(REPOXLS_BASE_PATH, PR_NAME, anio,  division)
        utlpy.create_dir_if_not_exists(REPOXLS_PATH)
        proc_genera_xls_by_div(division, anio)
